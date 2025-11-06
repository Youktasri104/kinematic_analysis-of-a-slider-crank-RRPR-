import numpy as np
from matplotlib import pyplot as plt
from matplotlib import animation
import  time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

#Inputs
a = float(input("Enter value of a : "))
l = float(input("Enter value of l : "))
theta_dot = float(input("Enter value of theta_dot : "))
theta_double_dot = float(input("Enter value of theta_double_dot : "))
theta1 = float(input("Enter value of theta : "))

#converting the theta to radians
theta1 = 180/np.pi

#offset value to a rod
delta = 5

#plotting
fig = plt.figure()
ax = plt.axes(xlim=(-(a+l+delta), a+l+delta), ylim=(-10, a+l+(3*delta)))
line, = ax.plot([], [], lw=1)

link,   = plt.plot([], [], 'b-', linewidth=4)
joints, = plt.plot([], [], 'ro', ls="", markersize=10)

#excel creation
file_name = 'new_data.xlsx'
try:
    workbook = load_workbook(file_name)
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook
    from openpyxl import Workbook
    workbook = Workbook()

new_sheet_name = f"Sheet_{len(workbook.sheetnames)}"
workbook.create_sheet(new_sheet_name)
worksheet = workbook[new_sheet_name]


row = 1
col = 1

worksheet.cell(row = row, column = col).value = "Time"
worksheet.cell(row = row, column = col+1).value = "theta"
worksheet.cell(row = row, column = col+2).value = "r"
worksheet.cell(row = row, column = col+3).value = "beta"
worksheet.cell(row = row, column = col+4).value = "r_dot"
worksheet.cell(row = row, column = col+5).value = "beta_dot"
worksheet.cell(row = row, column = col+6).value = "beta_double_dot"
worksheet.cell(row = row, column = col+7).value = "r_double_dot"
worksheet.cell(row = row, column = col+8).value = "Tangential Acc."
worksheet.cell(row = row, column = col+9).value = "Centrifugal Acc."
worksheet.cell(row = row, column = col+10).value = "Coriolis Acc."
worksheet.cell(row = row, column = col+11).value = "a"
worksheet.cell(row = row, column = col+12).value = "l"
worksheet.cell(row = row, column = col+13).value = "theta_dot"
worksheet.cell(row = row, column = col+14).value = "theta_double_dot"

bold_font = Font(bold=True)
center_alignment = Alignment(horizontal='center', vertical='center')
for cell in worksheet[1]:
    cell.font = bold_font
    cell.alignment = center_alignment

for column_cells in worksheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[column_cells[0].column_letter].width = 15

#finding the required values
p = l+(a*np.sin(theta1))
q = a*np.cos(theta1)

if(p>0 and q>0):
    beta1 = np.arctan(p/q)
if(p>0 and q<0):
    beta1 = np.pi + np.arctan(p/q)
if (p < 0 and q < 0):
    beta1 = 0 - np.pi + np.arctan(p / q)
if (p < 0 and q > 0):
    beta1 = np.arctan(p / q)

r1 = np.sqrt((a*a)+(l*l)+(2*a*l*np.sin(theta1)))

#Velocity analysis
r_dot = a*theta_dot*np.sin(beta1-theta1)
beta_dot = (a*theta_dot*np.cos(theta1-beta1))/r1

#Acceleration analysis
r_double_dot = (r1*beta_dot*beta_dot) - (a*theta_double_dot*np.sin(theta1-beta1))
beta_double_dot = -2*r_dot*beta_dot*((a * ((theta_double_dot * np.cos(theta1 - beta1) + (theta_dot * theta_dot * np.sin(theta1 - beta1))))) / r1)

print("For one specific value of theta")
print("beta = ", (beta1*180)/np.pi)
print("r = ", r1)
print("r_dot = ", r_dot)
print("beta_dot = ", beta_dot)
print("r_double_dot = ", r_double_dot)
print("beta_double_dot = ", beta_double_dot)

def get_values(theta, row):
    p = l+(a*np.sin(theta))
    q = a*np.cos(theta)

    if(p>0 and q>0):
        beta = np.arctan(p/q)
    if(p>0 and q<0):
        beta = np.pi + np.arctan(p/q)
    if (p < 0 and q < 0):
        beta = 0 - np.pi + np.arctan(p / q)
    if (p < 0 and q > 0):
        beta = np.arctan(p / q)

    r = np.sqrt((a*a)+(l*l)+(2*a*l*np.sin(theta)))

    # Velocity analysis
    r_dot = a * theta_dot * np.sin(beta - theta)
    beta_dot = (a * theta_dot * np.cos(theta - beta))/r

    # Acceleration analysis
    r_double_dot = (r * beta_dot * beta_dot) - (a * theta_double_dot * np.sin(theta - beta))
    beta_double_dot = -2*r_dot*beta_dot*((a * ((theta_double_dot * np.cos(theta - beta) + (theta_dot * theta_dot * np.sin(theta - beta))))) / r)

    t = time.time()
    times = datetime.fromtimestamp(t).strftime('%H:%M:%S ')

    col = 1

    worksheet.cell(row=row, column=col).value = times
    worksheet.cell(row=row, column=col+1).value = (theta*180)/np.pi
    worksheet.cell(row=row, column=col+2).value = r
    worksheet.cell(row=row, column=col+3).value = (beta*180)/np.pi
    worksheet.cell(row=row, column=col+4).value = r_dot
    worksheet.cell(row=row, column=col+5).value = beta_dot
    worksheet.cell(row=row, column=col+6).value = beta_double_dot
    worksheet.cell(row=row, column=col+7).value = r_double_dot
    worksheet.cell(row=row, column=col+8).value = r*theta_double_dot
    worksheet.cell(row=row, column=col+9).value = r*theta_dot*theta_dot
    worksheet.cell(row=row, column=col+10).value = 2*r*theta_dot
    worksheet.cell(row=row, column=col+11).value = a
    worksheet.cell(row=row, column=col+12).value = l
    worksheet.cell(row=row, column=col+13).value = theta_dot
    worksheet.cell(row=row, column=col+14).value = theta_double_dot

    A = [0,0]
    B = [0,l]
    C = [r*np.cos(beta), r*np.sin(beta)]
    D = [(l+a+delta)*np.cos(beta), (l+a+delta)*np.sin(beta)]

    return A, B, C, D

w = []
def animate(theta):
    w.append(theta)
    z = len(w)+1

    A, B, C, D = get_values(theta, z)

    x = [A[0], B[0], C[0], D[0], A[0]]
    y = [A[1], B[1], C[1], D[1], A[1]]

    link.set_data(x, y)
    joints.set_data(x, y)

    return link, joints


anim = animation.FuncAnimation(fig, animate, np.arange(0, 4*np.pi, 1), interval=100 ,blit=False)

plt.show()

workbook.save('new_data.xlsx')